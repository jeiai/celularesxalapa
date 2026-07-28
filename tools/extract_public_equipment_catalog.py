import argparse
import json
import re
from pathlib import Path

import openpyxl


DEFAULT_INPUT = Path(r"C:\Users\CORE\Documents\celularesxalapa\equipos v10.xlsx")
FALLBACK_INPUT = Path(r"C:\Users\CORE\Documents\celularesxalapa\V7.10    GENERAL .xlsx")
DEFAULT_OUTPUT = Path("public/data/equipos-catalogo.json")

ECONOMIC_TERMS = (
    "PRECIO",
    "PAGO",
    "MENSUAL",
    "ENGANCHE",
    "DIFERENCIA",
    "COSTO",
    "CARGO",
    "CASHBACK",
    "IVA",
)


def clean(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def normalize(value):
    return clean(value).upper()


def find_header(rows):
    for index, row in enumerate(rows[:15]):
        values = [normalize(value) for value in row]
        if "MARCA" in values and "MODELO" in values:
            return index
    return None


def get_cell(row, index):
    if index is None or index >= len(row):
        return ""
    return clean(row[index])


def is_economic_header(value):
    normalized = normalize(value)
    return any(term in normalized for term in ECONOMIC_TERMS)


def extract_items(path):
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    items = []
    seen = set()

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        header_index = find_header(rows)
        if header_index is None:
            continue

        header = [normalize(value) for value in rows[header_index]]
        header_lookup = {value: index for index, value in enumerate(header) if value and not is_economic_header(value)}

        category_index = header_lookup.get("CATEGORIA")
        brand_index = header_lookup.get("MARCA")
        model_index = header_lookup.get("MODELO")
        tech_index = header_lookup.get("TECNOLOGIA")
        color_index = header_lookup.get("COLOR") or header_lookup.get("COLORES")
        start_index = header_lookup.get("INICIO")
        end_index = header_lookup.get("FIN")

        for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            brand = get_cell(row, brand_index)
            model = get_cell(row, model_index)
            if not brand or not model or normalize(model) == "MODELO":
                continue

            category = get_cell(row, category_index) or sheet_name
            technology = get_cell(row, tech_index)
            color = get_cell(row, color_index)
            valid_from = get_cell(row, start_index)
            valid_to = get_cell(row, end_index)
            key = (normalize(brand), normalize(model))
            if key in seen:
                continue
            seen.add(key)

            items.append(
                {
                    "id": f"{sheet_name}-{row_number}".lower().replace(" ", "-"),
                    "sourceSheet": sheet_name,
                    "category": category,
                    "brand": brand,
                    "model": model,
                    "technology": technology,
                    "color": color,
                    "validFrom": valid_from,
                    "validTo": valid_to,
                }
            )

    items.sort(key=lambda item: (item["brand"], item["model"], item["category"]))
    return items


def main():
    parser = argparse.ArgumentParser(description="Extrae catalogo publico de equipos sin campos economicos.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    input_path = args.input
    source_note = "requested"
    if not input_path.exists() and input_path == DEFAULT_INPUT and FALLBACK_INPUT.exists():
        input_path = FALLBACK_INPUT
        source_note = "fallback"

    if not input_path.exists():
        raise FileNotFoundError(f"No se encontro el Excel: {args.input}")

    items = extract_items(input_path)
    payload = {
        "source": input_path.name,
        "sourceMode": source_note,
        "count": len(items),
        "filters": {
            "categories": sorted({item["category"] for item in items if item["category"]}),
            "brands": sorted({item["brand"] for item in items if item["brand"]}),
            "technologies": sorted({item["technology"] for item in items if item["technology"]}),
            "colors": sorted({item["color"] for item in items if item["color"]}),
        },
        "items": items,
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(json.dumps({"source": payload["source"], "sourceMode": source_note, "count": len(items)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
