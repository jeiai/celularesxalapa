import json
import os
import re
import sys
from pathlib import Path

import openpyxl


BASE_EXCLUDED = {
    "CATEGORIA",
    "MARCA",
    "MODELO",
    "TECNOLOGIA",
    "COLOR",
    "COLORES",
    "CLAVE SAP",
    "PREPAGO",
    "VIGENCIA",
    "INICIO",
    "FIN",
    "NOMBRECOMPRECIADOR",
    "CVEMODELOM2K",
}


def clean(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def as_number(value):
    if value in (None, "", "NA"):
        return None
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return None


def find_header(rows):
    for index, row in enumerate(rows[:12]):
        values = [clean(value).upper() for value in row]
        if "MARCA" in values and "MODELO" in values:
            return index
    return None


def header_map(row):
    return {clean(value).upper(): index for index, value in enumerate(row) if clean(value)}


def add_simple(items, rows, sheet, kind, header_index):
    headers = header_map(rows[header_index])
    for row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        category = clean(row[headers.get("CATEGORIA", 0)] if headers.get("CATEGORIA", 0) < len(row) else "")
        brand = clean(row[headers.get("MARCA", 1)] if headers.get("MARCA", 1) < len(row) else "")
        model = clean(row[headers.get("MODELO", 2)] if headers.get("MODELO", 2) < len(row) else "")
        if not model or model.upper() in {"MODELO", "EQUIPO CON SIM"}:
            continue
        items.append(
            {
                "id": f"{kind}-{sheet}-{row_index}".lower().replace(" ", "-"),
                "tipo": kind,
                "categoria": category,
                "marca": brand,
                "modelo": model,
                "tecnologia": clean(row[headers.get("TECNOLOGIA", 3)] if headers.get("TECNOLOGIA", 3) < len(row) else ""),
                "color": "",
                "precioConIva": as_number(row[headers.get("PRECIO PUBLICO PROMOCION CON IVA", 5)] if headers.get("PRECIO PUBLICO PROMOCION CON IVA", 5) < len(row) else None),
                "precioSinIva": as_number(row[headers.get("PRECIO PUBLICO PROMOCION SIN IVA", 4)] if headers.get("PRECIO PUBLICO PROMOCION SIN IVA", 4) < len(row) else None),
                "inicio": clean(row[headers.get("INICIO", 6)] if headers.get("INICIO", 6) < len(row) else ""),
                "fin": clean(row[headers.get("FIN", 7)] if headers.get("FIN", 7) < len(row) else ""),
                "planes": [],
            }
        )


def add_financed(items, rows, sheet, kind):
    header_index = find_header(rows)
    if header_index is None or header_index + 1 >= len(rows):
        return

    header = rows[header_index]
    subheader = rows[header_index + 1]
    headers = header_map(header)
    grouped = {}
    current_plan = ""

    for col_index, value in enumerate(header):
        top = clean(value)
        sub = clean(subheader[col_index] if col_index < len(subheader) else "")
        if top and top.upper() not in BASE_EXCLUDED:
            current_plan = top
        if current_plan and ("PRECIO" in sub.upper() or "PAGO MENSUAL" in sub.upper()):
            bucket = grouped.setdefault(current_plan, {})
            if "PAGO MENSUAL" in sub.upper():
                bucket["monthly"] = col_index
            else:
                bucket["price"] = col_index

    for row_index, row in enumerate(rows[header_index + 2 :], start=header_index + 3):
        brand = clean(row[headers.get("MARCA", 1)] if headers.get("MARCA", 1) < len(row) else "")
        model = clean(row[headers.get("MODELO", 2)] if headers.get("MODELO", 2) < len(row) else "")
        if not model:
            continue

        plans = []
        for plan, cols in grouped.items():
            price = as_number(row[cols["price"]] if cols.get("price") is not None and cols["price"] < len(row) else None)
            monthly = as_number(row[cols["monthly"]] if cols.get("monthly") is not None and cols["monthly"] < len(row) else None)
            if price is not None or monthly is not None:
                plans.append({"plan": plan, "precio": price, "mensualidad": monthly})

        if not plans:
            continue

        items.append(
            {
                "id": f"{kind}-{sheet}-{row_index}".lower().replace(" ", "-"),
                "tipo": kind,
                "categoria": clean(row[headers.get("CATEGORIA", 0)] if headers.get("CATEGORIA", 0) < len(row) else kind),
                "marca": brand,
                "modelo": model,
                "tecnologia": clean(row[headers.get("TECNOLOGIA", 3)] if headers.get("TECNOLOGIA", 3) < len(row) else ""),
                "color": clean(row[headers.get("COLOR", 4)] if headers.get("COLOR", 4) < len(row) else ""),
                "precioConIva": None,
                "precioSinIva": None,
                "inicio": "",
                "fin": "",
                "planes": plans,
            }
        )


def main():
    source = Path(sys.argv[1])
    output = Path(sys.argv[2])
    workbook = openpyxl.load_workbook(source, data_only=True, read_only=False)
    rows_by_sheet = {
        sheet: list(workbook[sheet].iter_rows(values_only=True))
        for sheet in workbook.sheetnames
    }

    items = []
    add_simple(items, rows_by_sheet["PREPAGO"], "PREPAGO", "Prepago", 4)
    add_simple(items, rows_by_sheet["ACCESORIOS - IOT"], "ACCESORIOS - IOT", "Accesorios", 0)

    for sheet, kind in [
        ("POSPAGO A PLAZO", "Postpago"),
        ("COMBOS_DUOS A PLAZO", "Combos"),
        ("EQUIPOS ZMA A PLAZO", "ZMA"),
        ("WIFI TELCEL A PLAZO", "WiFi Telcel"),
    ]:
        add_financed(items, rows_by_sheet[sheet], sheet, kind)

    payload = {
        "source": os.path.basename(source),
        "generatedFrom": "V7.10 GENERAL",
        "count": len(items),
        "filters": {
            "tipos": sorted({item["tipo"] for item in items}),
            "marcas": sorted({item["marca"] for item in items if item["marca"]}),
            "categorias": sorted({item["categoria"] for item in items if item["categoria"]}),
            "planes": sorted({plan["plan"] for item in items for plan in item["planes"]}),
        },
        "items": items,
    }

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(json.dumps({"count": len(items), "output": str(output), "types": payload["filters"]["tipos"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
